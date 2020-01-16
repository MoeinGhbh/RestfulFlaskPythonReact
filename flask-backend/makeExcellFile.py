import openpyxl
from xlrd import open_workbook
from xlsxwriter.utility import xl_rowcol_to_cell
from JsonUpdate import *


class makeExcellFile():

    @staticmethod
    def findCellDelete(zoneName, itemId):
        book = openpyxl.load_workbook('List.xlsx')
        for sheet in book:
            for row in sheet.iter_rows():
                flag = False
                for cell in row:
                    # if cell.value == -1:
                    #     break
                    if cell.value != None:
                        if sheet.cell(cell.row, 3).value == zoneName and sheet.cell(cell.row, 4).value == itemId:
                            sheet.cell(row=cell.row, column=1).value = None
                            sheet.cell(row=cell.row, column=2).value = None
                            sheet.cell(row=cell.row, column=3).value = None
                            sheet.cell(row=cell.row, column=4).value = None
                            sheet.cell(row=cell.row, column=5).value = None
                            sheet.cell(row=cell.row, column=6).value = None
                            sheet.cell(row=cell.row, column=7).value = None
                            book.save('List.xlsx')
                    #         flag = True
                    # if flag:
                    #     break
                    # else:
                    #     continue

    @staticmethod
    def findCell(mySheet, zoneName, itemId):
        book = open_workbook('List.xlsx')
        for sheet in book.sheets():
            if sheet.name == mySheet:
                for rowidx in range(sheet.nrows):
                    row = sheet.row(rowidx)
                    for cell in enumerate(row):
                        if sheet.cell(rowidx, 2).value == zoneName and sheet.cell(rowidx, 3).value == itemId:
                            return False
        return True

    @staticmethod
    def makeFile(zoneId, zoneName, newitems):
        wb = openpyxl.load_workbook('List.xlsx')
        HighAmper = wb['HighAmper']
        for item in newitems:
            if makeExcellFile.findCell('HighAmper', zoneName, item['itemId']):
                if (item['group'] == 'Aircondition') and (
                        (item['speedType'] == 'Fast') or (item['speedType'] == 'Normal')):
                    flag = False
                    for row in HighAmper.iter_rows():
                        for cell in row:
                            if cell.value == None:
                                HighAmper.cell(row=cell.row, column=1).value = (cell.row - 1)
                                HighAmper.cell(row=cell.row, column=2).value = zoneId
                                HighAmper.cell(row=cell.row, column=3).value = zoneName
                                HighAmper.cell(row=cell.row, column=4).value = item['itemId']
                                HighAmper.cell(row=cell.row, column=5).value = item['itemName']
                                HighAmper.cell(row=cell.row, column=6).value = item['group']
                                HighAmper.cell(row=cell.row, column=7).value = item['speedType']
                                JsonUpdate.UpdateStaticCode(zoneId, item['itemId'], (cell.row - 1))
                                flag = True
                                break
                        if flag:
                            break
                        else:
                            continue
            wb.save('List.xlsx')

        wb = openpyxl.load_workbook('List.xlsx')
        LowAmper = wb['LowAmper']
        for item in newitems:
            if makeExcellFile.findCell('LowAmper', zoneName, item['itemId']):
                if item['group'] != 'Aircondition' or (
                        (item['group'] == 'Aircondition') and (item['speedType'] == 'Slow')):
                    flag = False
                    for row in LowAmper.iter_rows():
                        for cell in row:
                            if cell.value == None and cell.row > 17:
                                LowAmper.cell(row=cell.row, column=1).value = (cell.row - 1) + 17
                                LowAmper.cell(row=cell.row, column=2).value = zoneId
                                LowAmper.cell(row=cell.row, column=3).value = zoneName
                                LowAmper.cell(row=cell.row, column=4).value = item['itemId']
                                LowAmper.cell(row=cell.row, column=5).value = item['itemName']
                                LowAmper.cell(row=cell.row, column=6).value = item['group']
                                if item['group'] == 'Aircondition':
                                    LowAmper.cell(row=cell.row, column=7).value = item['speedType']
                                else:
                                    LowAmper.cell(row=cell.row, column=7).value = 'null'
                                JsonUpdate.UpdateStaticCode(zoneId, item['itemId'], (cell.row - 1))
                                flag = True
                                break
                        if flag:
                            break
                        else:
                            continue
            wb.save('List.xlsx')
